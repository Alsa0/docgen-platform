import os
import shutil
import copy
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from config.settings import OUTPUT_DIR, TEMPLATES_DIR

def copy_cell_style(source_cell, target_cell):
    """Copie le style complet d'une cellule vers une autre en gérant les MergedCells."""
    if not source_cell.has_style:
        return
    
    # On cherche la cellule cible réelle si c'est une MergedCell
    actual_target = target_cell
    if target_cell.__class__.__name__ == 'MergedCell':
        # On ne peut pas appliquer de style individuel à une MergedCell secondaire
        # Le style doit être appliqué à la Master Cell. 
        # Mais ici, on suppose que target_cell est déjà géré ou on l'ignore.
        return

    actual_target.font = copy.copy(source_cell.font)
    actual_target.border = copy.copy(source_cell.border)
    actual_target.fill = copy.copy(source_cell.fill)
    actual_target.number_format = source_cell.number_format
    actual_target.protection = copy.copy(source_cell.protection)
    actual_target.alignment = copy.copy(source_cell.alignment)

def apply_magenta_style(cell, bold=False, align='center', font_size=10):
    """Applique le style magenta (bordures et police) aux cellules."""
    magenta = "B01B6B"
    cell.font = Font(name='Calibri', size=font_size, bold=bold, color=magenta)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    
    magenta_side = Side(style='thin', color=magenta)
    cell.border = Border(left=magenta_side, right=magenta_side, top=magenta_side, bottom=magenta_side)

def safe_set(ws, row, col, value=None, font=None, number_format=None, alignment=None, border=None, fill=None):
    """Définit les propriétés d'une cellule en gérant les fusions (Master vs Merged)."""
    cell = ws.cell(row=row, column=col)
    target = cell
    
    # Si c'est une cellule fusionnée (secondaire), on cherche la cellule maîtresse
    if cell.__class__.__name__ == 'MergedCell':
        coord = cell.coordinate
        for merged_range in ws.merged_cells.ranges:
            if coord in merged_range:
                target = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                break
    
    # On applique les propriétés à la cible (Master cell ou cellule normale)
    if value is not None:
        target.value = value
    if font is not None:
        target.font = font
    if number_format is not None:
        target.number_format = number_format
    if alignment is not None:
        target.alignment = alignment
    if border is not None:
        target.border = border
    if fill is not None:
        target.fill = fill

def generate_excel_bom(config, bom_items):
    """
    Remplit le template Excel BOM_1.xlsx.
    Mapping : Code (B) = Marque/Modèle, Description (D) = Désignation, Unit (H) = "Unit", ...
    """
    template_path = os.path.join(TEMPLATES_DIR, "BOM_1.xlsx")
    filename = f"BOM_{config.get('client_nom', 'client')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    
    if not os.path.exists(template_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        return generate_bom_fallback(config, bom_items, ws, output_path)

    shutil.copy(template_path, output_path)
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active
    
    # 1. Infos projet
    client_name = str(config.get('client_nom', 'N/A')).upper()
    safe_set(ws, 2, 2, value=f"Client : {client_name}")
    
    type_projet = config.get('type_projet', 'Équipements')
    title = f"Équipement {type_projet}" if "équipement" not in type_projet.lower() else type_projet
    safe_set(ws, 3, 2, value=title)
    
    # 2. Localisation et Nettoyage
    start_row = 4
    total_row = None
    for r in range(start_row, 100):
        for c in range(1, 12):
            val = str(ws.cell(row=r, column=c).value or "").upper()
            if "TOTAL" in val:
                total_row = r
                break
        if total_row: break
    
    if not total_row: total_row = 10
    
    # --- MÉTHODE RADICALE : Dé-fusionner toute la zone de données ---
    data_end_search = max(total_row + 50, 100)
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if merged_range.min_row >= start_row and merged_range.max_row <= data_end_search:
            if merged_range.min_col >= 1 and merged_range.max_col <= 11:
                try:
                    ws.unmerge_cells(str(merged_range))
                except: pass

    style_ref_row = start_row
    
    # Vider les colonnes
    for r in range(start_row, total_row):
        for c in range(1, 12):
            ws.cell(row=r, column=c).value = None

    num_samples = total_row - start_row
    num_items = len(bom_items)
    
    if num_items > num_samples:
        ws.insert_rows(total_row, num_items - num_samples)
    elif num_items < num_samples:
        ws.delete_rows(start_row + num_items, num_samples - num_items)
    
    new_total_row = start_row + num_items

    # 3. Remplissage des articles
    for i, item in enumerate(bom_items):
        row = start_row + i
        
        ws.cell(row=row, column=1).value = f"1.{i+1}"
        marque = item.get('marque', '')
        modele = item.get('modele', '')
        ws.cell(row=row, column=2).value = f"{marque} {modele}".strip()
        
        desc_val = item.get('designation', '') or item.get('description', '')
        ws.cell(row=row, column=4).value = desc_val
        
        ws.cell(row=row, column=8).value = "Unit"
        ws.cell(row=row, column=9).value = item.get('quantite', 1)
        ws.cell(row=row, column=10).value = item.get('prix_unitaire', 0)
        
        qty = item.get('quantite', 1)
        u_p = item.get('prix_unitaire', 0)
        ws.cell(row=row, column=11).value = item.get('prix_total', qty * u_p)
        
        # RE-FUSIONNER
        try:
            ws.merge_cells(start_row=row, end_row=row, start_column=2, end_column=3)
            ws.merge_cells(start_row=row, end_row=row, start_column=4, end_column=7)
        except: pass
        
        # Styles
        for c in range(1, 12):
            copy_cell_style(ws.cell(row=style_ref_row, column=c), ws.cell(row=row, column=c))

    # 4. Total Global
    # On recalcule systématiquement à partir des items pour être sûr que ça corresponde au tableau web
    total_val = sum(item.get('prix_total', 0) for item in bom_items)
    
    # Bordure Magenta pour la ligne de total
    magenta = "B01B6B"
    side = Side(style='thin', color=magenta)
    total_border = Border(top=side, bottom=side)
    
    # On place le libellé en colonne J (10)
    safe_set(ws, new_total_row, 10, value="SUB-TOTAL Equipements", 
             font=Font(bold=True, color=magenta),
             alignment=Alignment(horizontal='right'),
             border=total_border)
    
    # On place le montant en colonne K (11)
    safe_set(ws, new_total_row, 11, value=total_val, 
             font=Font(bold=True), 
             number_format='"Ar" #,##0.00',
             border=total_border)
    
    # Ajuster la largeur de la colonne K (Amount) pour éviter les #######
    ws.column_dimensions['K'].width = 18
    
    wb.save(output_path)
    return output_path

def generate_bom_fallback(config, bom_items, ws, output_path):
    """Ancienne méthode de génération de zéro si le template est manquant."""
    headers = ["#", "Code", "Description", "Unit", "Qty", "Unit Price", "Amount"]
    magenta = "B01B6B"
    header_font = Font(name='Calibri', size=11, bold=True, color=magenta)
    
    for col, text in enumerate(headers, 1):
        safe_set(ws, 1, col, value=text, font=header_font)

    wb = ws.parent
    wb.save(output_path)
    return output_path

def generate_excel_sow(config, sow_content, bom_items=None):
    """Génère un fichier Excel SOW."""
    template_path = os.path.join(TEMPLATES_DIR, "SOW_1.xlsx")
    filename = f"SOW_{config.get('client_nom', 'client')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)
    
    if not os.path.exists(template_path):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SoW"
    else:
        shutil.copy(template_path, output_path)
        wb = openpyxl.load_workbook(output_path)
        ws = wb["SoW"] if "SoW" in wb.sheetnames else wb.active
    
    safe_set(ws, 10, 4, value=config.get('client_nom', 'N/A'), font=Font(bold=True, color="B01B6B"))
    safe_set(ws, 12, 4, value=config.get('site_intervention', 'Antananarivo'))
    safe_set(ws, 13, 4, value=config.get('projet_description', ''))
    
    taches = sow_content.get('taches', [])
    start_row = 20
    
    for i, t in enumerate(taches):
        row = start_row + i
        if i > 0: ws.insert_rows(row)
        
        safe_set(ws, row, 2, value=f"Phase {t.get('phase_num', 1)}")
        safe_set(ws, row, 3, value=t.get('type', 'Implémentation'))
        safe_set(ws, row, 4, value=t.get('titre', ''))
        safe_set(ws, row, 5, value=t.get('description', ''))
        safe_set(ws, row, 6, value=config.get('site_intervention', 'Client Site'))
        safe_set(ws, row, 7, value="Ing-1")
        safe_set(ws, row, 8, value=1)
        safe_set(ws, row, 9, value=t.get('duree_jours', 1))

    wb.save(output_path)
    return output_path
